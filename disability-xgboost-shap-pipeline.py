# -*- coding:gb2312 -*-

import os
from datetime import datetime
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import shap
from xgboost import XGBClassifier
from sklearn.model_selection import train_test_split, StratifiedKFold
from sklearn.metrics import (roc_auc_score, accuracy_score, precision_score, 
                             recall_score, f1_score, confusion_matrix, 
                             classification_report, roc_curve, brier_score_loss)
from openpyxl import Workbook
from scipy.stats import norm
from sklearn.calibration import calibration_curve
from statsmodels.nonparametric.smoothers_lowess import lowess
from sklearn.utils import resample
from sklearn.base import clone
from matplotlib.ticker import MaxNLocator

def create_results_directory(base_dir="results"):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    results_dir = os.path.join(base_dir, f"results_{timestamp}")
    os.makedirs(results_dir, exist_ok=True)
    return results_dir

def encode_features(X, categorical_columns):
    X_encoded = X.copy()
    for col in categorical_columns:
        X_encoded[col] = X_encoded[col].astype('category').cat.codes
    return X_encoded

def split_data(X, y, test_size=0.2):
    return train_test_split(X, y, test_size=test_size, stratify=y, random_state=42)

def train_xgb_model(X_train, y_train):
    model = XGBClassifier(
        objective='binary:logistic',
        eval_metric='aucpr',
        booster='gbtree',
        tree_method='auto',
        scale_pos_weight=4,
        learning_rate=0.014465,
        min_child_weight=4,
        subsample=0.715168,
        colsample_bytree=0.741349,
        max_delta_step=3,
        gamma=0.318761,
        max_depth=3,
        n_estimators=1300,
        reg_lambda=0.589819,
        reg_alpha=4.537696,
        use_label_encoder=False,
        random_state=42,
        n_jobs=-1
    )
    y_train = y_train.astype(int)
    model.fit(X_train, y_train)
    return model

def plot_confusion_matrix(cm, labels_name, title, save_path):
    plt.figure()
    plt.matshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
    for i in range(len(cm)):
        for j in range(len(cm)):
            plt.annotate(cm[j, i], xy=(i, j), horizontalalignment='center', verticalalignment='center',
                         bbox=dict(boxstyle='round', pad=0.5, fc='w', ec='k', lw=1, alpha=0.5))
    plt.xticks(range(len(labels_name)), labels_name)
    plt.yticks(range(len(labels_name)), labels_name)
    plt.title(title)
    plt.ylabel('True label')
    plt.xlabel('Predicted label')
    plt.colorbar()
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()

def AUC_CI(auc, y_true, alpha=0.05):
    y_true = np.array(y_true)
    n1, n2 = np.sum(y_true == 1), np.sum(y_true == 0)
    q1 = auc / (2 - auc)
    q2 = (2 * auc ** 2) / (1 + auc)
    se = np.sqrt((auc*(1-auc) + (n1-1)*(q1-auc**2) + (n2-1)*(q2-auc**2)) / (n1*n2))
    z_lower, z_upper = norm.ppf([alpha/2, 1-alpha/2])
    lowerb, upperb = auc + z_lower*se, auc + z_upper*se
    return lowerb, upperb

from sklearn.calibration import calibration_curve

def evaluate_model(model, X_test, y_test, results_dir, prefix, threshold=0.540, n_bins=10):
    y_pred_proba = model.predict_proba(X_test)[:,1]
    y_pred = (y_pred_proba >= threshold).astype(int)
    y_test = y_test.astype(int)

    # -------------------------
    # confusion_matrix
    # -------------------------
    cm = confusion_matrix(y_test, y_pred)
    plt.figure()
    plt.matshow(cm, interpolation='nearest', cmap=plt.cm.Blues)

    plt.gca().set_xticks(np.arange(-0.5, len(cm), 1), minor=True)
    plt.gca().set_yticks(np.arange(-0.5, len(cm), 1), minor=True)
    plt.grid(which='minor', color='lightgray', linestyle='-', linewidth=1)
    plt.gca().tick_params(which='minor', bottom=False, left=False)

    for i in range(len(cm)):
        for j in range(len(cm)):
            plt.annotate(cm[j, i], xy=(i, j), 
                         horizontalalignment='center', verticalalignment='center',
                         bbox=dict(boxstyle='round', pad=0.5, fc='w', ec='k', lw=1, alpha=0.5))

    plt.xticks(range(len(['0','1'])), ['0','1'])
    plt.yticks(range(len(['0','1'])), ['0','1'])
    plt.title(f'Confusion Matrix')
    plt.ylabel('True label')
    plt.xlabel('Predicted label')
    plt.colorbar()
    plt.savefig(os.path.join(results_dir, f'{prefix}_confusion_matrix.tif'), dpi=300, bbox_inches='tight')
    plt.close()

    # -------------------------
    # ROC curve
    # -------------------------
    fpr, tpr, thresholds = roc_curve(y_test, y_pred_proba)
    auc_score = roc_auc_score(y_test, y_pred_proba)
    auc_lower, auc_upper = AUC_CI(auc_score, y_test)
    plt.figure()
    plt.plot(fpr, tpr, color='crimson', lw=2, label=f'ROC curve (AUC={auc_score:.4f})')
    plt.fill_between(fpr, tpr, alpha=0.2, color='lightsteelblue', 
                     label=f'AUC 95%CI: ({auc_lower:.4f}, {auc_upper:.4f})')
    plt.plot([-0.05,1.05], [-0.05,1.05], linestyle='--')
    plt.xlim([-0.05,1.05]); plt.ylim([-0.05,1.05])
    plt.xlabel('False Positive Rate'); plt.ylabel('True Positive Rate')
    plt.title(f'ROC Curve'); plt.legend(loc='lower right')
    plt.savefig(os.path.join(results_dir, f'{prefix}_ROC.tif'), dpi=300, bbox_inches='tight')
    plt.close()

    prob_true, prob_pred = calibration_curve(y_test, y_pred_proba, n_bins=n_bins, strategy='uniform')
    
    fig, ax1 = plt.subplots(figsize=(6,6))

    # Calibration curve
    ax1.plot(prob_pred, prob_true, marker='D', markersize=4, linewidth=2, color='#1f77b4', label='Calibration curve')
    ax1.plot([0,1], [0,1], linestyle='--', color='gray', label='Perfectly calibrated')
    ax1.set_xlabel('Predicted probability')
    ax1.set_ylabel('Observed frequency', color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    ax1.set_xlim(0,1)
    ax1.set_ylim(0,1)

    ax2 = ax1.twinx()
    counts, bins, patches = ax2.hist(
        y_pred_proba, bins=20, range=(0,1), color='skyblue', alpha=0.3,
        edgecolor='steelblue', linewidth=1, density=True, label='Probability distribution'
    )
    ax2.set_ylabel('Density', color='steelblue')
    ax2.tick_params(axis='y', labelcolor='steelblue')
    ax2.set_ylim(0, max(counts)*1.1)

    # legend
    lines_1, labels_1 = ax1.get_legend_handles_labels()
    lines_2, labels_2 = ax2.get_legend_handles_labels()
    ax1.legend(lines_1 + lines_2, labels_1 + labels_2, loc='upper right')

    #plt.title('Calibration Curve with Probability Distribution')
    plt.tight_layout()
    plt.savefig(os.path.join(results_dir, f'{prefix}_calibration_curve.tif'), dpi=300, bbox_inches='tight')
    plt.close()
    # -------------------------
    # save metrics to Excel
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws['A1'] = auc_score
    ws['A2'] = f'[{auc_lower:.4f},{auc_upper:.4f}]'
    ws['A3'] = accuracy_score(y_test, y_pred)
    ws['A4'] = precision_score(y_test, y_pred, average='weighted')
    ws['A5'] = recall_score(y_test, y_pred, average='weighted')
    ws['A6'] = f1_score(y_test, y_pred, average='weighted')
    report = classification_report(y_test, y_pred, output_dict=True)
    report_str = ""
    for label, metrics in report.items():
        if isinstance(metrics, dict):
            report_str += f"{label}: "
            report_str += ", ".join([f"{k}={v:.4f}" for k, v in metrics.items()])
            report_str += "\n"
        else:
            report_str += f"{label}: {metrics:.4f}\n"
    ws['A7'] = report_str
    wb.save(os.path.join(results_dir, f'{prefix}_metrics.xlsx'))

def shap_analysis(model, X_test, results_dir, prefix):
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X_test)

    # waterfall plot/decision plot
    idx_samples = [0, 20, 154, 999] 
    for i, idx in enumerate(idx_samples, start=1):
        shap.waterfall_plot(shap.Explanation(shap_values[idx], explainer.expected_value, X_test.iloc[idx]), show=False)
        plt.savefig(os.path.join(results_dir, f'{prefix}_Figure_{i}_waterfall.tif'), dpi=300, bbox_inches='tight')
        plt.close()
        shap.decision_plot(explainer.expected_value, shap_values[idx], X_test.iloc[idx], show=False)
        plt.savefig(os.path.join(results_dir, f'{prefix}_Figure_{i}_decision.tif'), dpi=300, bbox_inches='tight')
        plt.close()

    # bar plot and summary plot
    shap.summary_plot(shap_values, X_test, max_display=25, show=False)
    plt.savefig(os.path.join(results_dir, f'{prefix}_Figure_summary_plot.tif'), dpi=300, bbox_inches='tight')
    plt.close()

    shap.plots.bar(explainer(X_test), max_display=25, show=False)
    plt.savefig(os.path.join(results_dir, f'{prefix}_Figure_bar_plot.tif'), dpi=300, bbox_inches='tight')
    plt.close()


def detect_plateau(feature_vals, shap_values, tolerance=0.1):
    plateau_start = None
    unique_vals = np.sort(np.unique(feature_vals))
    mean_shap = [shap_values[feature_vals == v].mean() for v in unique_vals]
    
    for i in range(1, len(mean_shap)):
        change = abs(mean_shap[i] - mean_shap[i-1])
        if change < tolerance:
            plateau_start = unique_vals[i]
            break
    if plateau_start is None:
        plateau_start = unique_vals[-1]
    return plateau_start, unique_vals, mean_shap

def plot_partial_dependence_auto_plateau(feature_name, X_test_encoded, X_test_raw, shap_values, results_dir, prefix):
    feature_vals = X_test_raw[feature_name].astype(int).values
    feature_shap = shap_values[:, X_test_encoded.columns.get_loc(feature_name)]

    # auto detect plateau
    plateau_threshold, unique_vals, mean_shap = detect_plateau(feature_vals, feature_shap)

    labels_map = {1:"very poor (1)", 2:"poor (2)", 3:"fair (3)", 4:"good (4)", 5:"very good (5)"}

    plateau_mask = feature_vals >= plateau_threshold
    plateau_rate = np.sum(plateau_mask) / len(feature_vals)

    fig, ax = plt.subplots(figsize=(7,4))
    ax.scatter(feature_vals, feature_shap, alpha=0.6, color='dodgerblue', label='SHAP value')

    ax.fill_betweenx(
        y=[min(feature_shap), max(feature_shap)],
        x1=plateau_threshold - 0.5,
        x2=max(feature_vals) + 0.5,
        color='lightcoral', alpha=0.3,
        label=f'Plateau region ({plateau_rate:.1%} samples)'
    )

    ax.axvline(plateau_threshold, color='red', linestyle='--', 
               label=f'Plateau threshold = {plateau_threshold}')

    ax.set_xticks(list(labels_map.keys()))
    ax.set_xticklabels([labels_map[i] for i in labels_map.keys()])

    ax.set_xlabel(feature_name)
    ax.set_ylabel('SHAP value')
    ax.set_title(f'Partial Dependence of {feature_name}')
    ax.legend(loc='upper right')
    plt.tight_layout()

    save_path = os.path.join(results_dir, f'{prefix}_{feature_name}_partial_dependence.tif')
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()

    return plateau_threshold, plateau_rate, unique_vals, mean_shap

def shap_bootstrap_analysis(X, y, params, results_dir, prefix,
                            n_bootstrap=100, sample_frac=0.8):

    feature_names = X.columns
    all_results = []

    X_test_fixed = X.copy()
    y_test_fixed = y.copy()

    for i in range(n_bootstrap):

        # bootstrap sampling
        X_resampled, y_resampled = resample(
            X, y,
            replace=True,
            n_samples=int(len(X) * sample_frac),
            random_state=42 + i
        )

        model = XGBClassifier(**params)
        model.fit(X_resampled, y_resampled)

        explainer = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X_test_fixed)

        mean_shap = np.abs(shap_values).mean(axis=0)

        all_results.append(mean_shap)

    all_results = np.array(all_results)

    df = pd.DataFrame({
        "feature": feature_names,
        "mean_shap": all_results.mean(axis=0),
        "std_shap": all_results.std(axis=0),
        "ci_low": np.percentile(all_results, 2.5, axis=0),
        "ci_high": np.percentile(all_results, 97.5, axis=0)
    })

    df = df.sort_values("mean_shap", ascending=False)

    df.to_csv(os.path.join(results_dir, f"{prefix}_shap_bootstrap.csv"), index=False)

    return df

def plot_bootstrap_shap(df, top_k=20):

    df_plot = df.head(top_k)

    plt.figure(figsize=(8,6))

    x = range(len(df_plot))

    # mean
    plt.errorbar(
        x,
        df_plot["mean_shap"],
        yerr=[
            df_plot["mean_shap"] - df_plot["ci_low"],
            df_plot["ci_high"] - df_plot["mean_shap"]
        ],
        fmt='o',
        capsize=4
    )

    for i, v in enumerate(df_plot["mean_shap"]):
        plt.text(
            i,
            v + 0.01,  
            f"{v:.2f}",
            ha='center',
            va='bottom',
            fontsize=8,
            bbox=dict(facecolor='white', edgecolor='none', alpha=0.7)
        )

    plt.xticks(x, df_plot["feature"], rotation=60, ha='right')
    plt.ylabel("Mean |SHAP value|")
    plt.tight_layout()
    plt.savefig(
        os.path.join(results_dir, "shap_bootstrap.tif"),
        dpi=300,
        bbox_inches='tight'
    )

def shap_feature_selection(X_train, X_test, y_train, y_test, model_params, results_dir):

    features = list(X_train.columns)
    results = []

    while len(features) > 0:

        model = XGBClassifier(**model_params)
        model.fit(X_train[features], y_train)

        y_pred = model.predict_proba(X_test[features])[:,1]
        auc = roc_auc_score(y_test, y_pred)

        explainer = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X_train[features])

        importance = np.abs(shap_values).mean(axis=0)

        worst_feature = features[np.argmin(importance)]

        results.append({
            "n_features": len(features),
            "features": ",".join(features),  
            "auc": auc,
            "removed": worst_feature
        })

        features.remove(worst_feature)

    df = pd.DataFrame(results)

    df.to_excel(
        os.path.join(results_dir, "feature_selection_path.xlsx"),
        index=False
    )

    return df
    
    
def plot_auc_drop(df, results_dir, prefix):
    plt.figure()

    plt.plot(df["n_features"], df["auc"], marker='o')

    plt.xlabel("Number of Features")
    plt.ylabel("AUC")

    plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    plt.gca().invert_xaxis() 

    plt.tight_layout()
    plt.savefig(os.path.join(results_dir, f"{prefix}_auc_drop_curve.tif"),
                dpi=300, bbox_inches='tight')
    plt.close()

# -----------------------------
# main
# -----------------------------
if __name__ == "__main__":
    datasets = {
        "all": ['adl','gender','residence_type','living_with_spouse','living_alone','social_pension','medical_insurance','self_reported_health',
                'physical_examination','falls','body_pain','sleep_duration','nap','sports_level','socialization','smoke','drink','depression',
                'living_with_child','financial_support_from_child','work'],
        "micro": ['adl','gender','self_reported_health','physical_examination','falls','body_pain','sleep_duration','nap','sports_level','smoke','drink','depression'],
        "meso": ['adl','living_with_spouse','living_alone','socialization','living_with_child','financial_support_from_child'],
        "macro": ['adl','residence_type','social_pension','medical_insurance']
    }

    data_files = {
        "all": "adl_status_calculation_9_11.csv",
        "micro": "adl_status_calculation_9_11_micro.csv",
        "meso": "adl_status_calculation_9_11_meso.csv",
        "macro": "adl_status_calculation_9_11_macro.csv"
    }

    base_results_dir = create_results_directory()

    for prefix, file_path in data_files.items():
        print(f"Processing dataset: {prefix}")
        results_dir = os.path.join(base_results_dir, prefix)
        os.makedirs(results_dir, exist_ok=True)

        train = pd.read_csv(file_path)
        categorical_cols = datasets[prefix]
        target_col = 'adl'
        categorical_cols = [col for col in categorical_cols if col != target_col]
        print([col for col in categorical_cols if col not in train.columns])
        
        train[categorical_cols] = train[categorical_cols].astype(str)
        X_raw = train[[col for col in train.columns if col != target_col]].copy()  
        
        non_feature_cols = ['adl', 'id']  
        features = [x for x in train.columns if x not in non_feature_cols]
        X = train[features]
        y = train['adl']

        # encode features
        X_encoded = encode_features(X, categorical_cols)

        # data split
        X_train, X_test, y_train, y_test = split_data(X_encoded, y)
        X_train_raw, X_test_raw, _, _ = split_data(X_raw, y) 

        X_train.assign(label=y_train).to_csv(os.path.join(results_dir, f'{prefix}_train.csv'), index=False)
        X_test.assign(label=y_test).to_csv(os.path.join(results_dir, f'{prefix}_test.csv'), index=False)

        # train xgboost model
        model = train_xgb_model(X_train, y_train)

        # model evaluate
        evaluate_model(model, X_test, y_test, results_dir, prefix, threshold=0.540, n_bins=10)
        
        # calibration_curve
        #plot_calibration_curve(model, X_test, y_test, results_dir, prefix)
        

        # SHAP analysis（base test set）
        shap_analysis(model, X_test, results_dir, prefix)
        

        # partial dependence + plateau
        # -------------------------------
        if prefix == "all":
            feature_name = 'self_reported_health' 
            X_test_raw = X_test.copy()  
            X_test_raw[feature_name] = train.loc[X_test.index, feature_name] 

            feature_name = 'self_reported_health'
            plateau_threshold, plateau_rate, unique_vals, mean_shap = \
                plot_partial_dependence_auto_plateau(
                    feature_name=feature_name,
                    X_test_encoded=X_test,
                    X_test_raw=X_test_raw,
                    shap_values=shap.TreeExplainer(model).shap_values(X_test),
                    results_dir=results_dir,
                    prefix=prefix
                )
            print(f"[{prefix}] {feature_name} plateau threshold: {plateau_threshold}, plateau rate: {plateau_rate:.2%}")
            
            # ========== SHAP robustness analysis ==========
            xgb_params = {
                "objective": "binary:logistic",
                "eval_metric": "aucpr",
                "booster": "gbtree",
                "tree_method": "auto",
                "scale_pos_weight": 4,
                "learning_rate": 0.014465,
                "min_child_weight": 4,
                "subsample": 0.715168,
                "colsample_bytree": 0.741349,
                "max_delta_step": 3,
                "gamma": 0.318761,
                "max_depth": 3,
                "n_estimators": 1300,
                "reg_lambda": 0.589819,
                "reg_alpha": 4.537696,
                "use_label_encoder": False,
                "random_state": 42,
                "n_jobs": -1}
            shap_bootstrap_df = shap_bootstrap_analysis(
                X=X_encoded,
                y=y.astype(int),
                params=xgb_params,
                results_dir=results_dir,
                prefix=prefix,
                n_bootstrap=100,
                sample_frac=0.8
            )
            plot_bootstrap_shap(shap_bootstrap_df, top_k=20)
            
            # ========== FEATURE SELECTION + AUC DROP ==========
            selection_df = shap_feature_selection(
                X_train, X_test, y_train, y_test,
                model_params=xgb_params,
                results_dir=results_dir  
            )
            plot_auc_drop(selection_df, results_dir, prefix)
               
    print("All datasets processed successfully!")
